# -*- coding: utf-8 -*-
"""
Sabake_osakana 仕様書Excel生成スクリプト
動的に更新可能な形式

使い方:
  python sabake_spec_generator.py

出力先:
  G:/マイドライブ/t0kag3/AIJiyukenkyu/Game/Unity/Sabake_osakana/docs/specification.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

# ============================================
# 設定
# ============================================
OUTPUT_PATH = "/mnt/c/temp/specification.xlsx"
ASSET_BASE = "G:/マイドライブ/t0kag3/AIJiyukenkyu/Game/Unity/Sabake_osakana/ForlocalAsset"

# スタイル定義
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_header_style(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

def apply_section_style(cell):
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.border = THIN_BORDER

def apply_normal_style(cell):
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical='center', wrap_text=True)

def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

# ============================================
# シート1: 全体概要
# ============================================
def create_overview_sheet(wb):
    ws = wb.active
    ws.title = "01_全体概要"

    # ヘッダー
    headers = ["セクション", "項目", "値", "備考"]
    set_column_widths(ws, [15, 25, 40, 40])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # データ定義（編集しやすい形式）
    data = [
        # セクション, 項目, 値, 備考
        ("基本情報", "プロジェクト名", "Sabake_osakana", "魚さばきブロック崩し"),
        ("", "ジャンル", "パズル/アクション", ""),
        ("", "プラットフォーム", "PC/スマホ/タブレット", ""),
        ("", "Unity版", "6000.2.14f1", ""),
        ("", "レンダリング", "Built-in 2D", ""),
        ("", "", "", ""),
        ("コンセプト", "ゲーム概要", "ブロックを崩して魚をさばく", "脱衣ブロック崩しの魚版"),
        ("", "目標", "全ブロックを崩して骨を露出させる", ""),
        ("", "演出", "魚→骨へ変化", "将来的に魚→切り身→骨の3段階"),
        ("", "", "", ""),
        ("レイヤー構造", "Layer1（最背面）", "骨の絵", "sakana_bone.png（要作成）"),
        ("", "Layer2（前面）", "魚の絵", "sakana_normal.png（既存）"),
        ("", "Layer3（最前面）", "グリッドブロック", "プログラム生成（グレー線）"),
        ("", "将来拡張", "切り身レイヤー追加", "sakana_sliced.png"),
        ("", "", "", ""),
        ("ブロック仕様", "表示方式", "魚画像をクリッピング表示", "方眼紙スタイル"),
        ("", "枠線", "グレー", "手描き風の質感"),
        ("", "行数", "5", "調整可能"),
        ("", "列数", "8", "調整可能"),
        ("", "合計", "40", "5×8"),
        ("", "", "", ""),
        ("操作方法", "PC（キーボード）", "←→キーでパドル移動", ""),
        ("", "PC（キーボード）", "スペースキーでボール発射", ""),
        ("", "PC（マウス）", "カーソル追従でパドル移動", ""),
        ("", "PC（マウス）", "クリックでボール発射", ""),
        ("", "モバイル", "タッチ追従でパドル移動", ""),
        ("", "モバイル", "タップでボール発射", ""),
        ("", "", "", ""),
        ("スコア", "ブロック破壊", "10点/個", ""),
        ("", "残機", "3", "初期値"),
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_data[0] and col_idx == 1:  # セクション名
                apply_section_style(cell)
            else:
                apply_normal_style(cell)

    return ws

# ============================================
# シート2: ファイル依存関係
# ============================================
def create_dependencies_sheet(wb):
    ws = wb.create_sheet("02_ファイル依存関係")

    headers = ["ファイル名", "種別", "場所", "依存先", "状態", "備考"]
    set_column_widths(ws, [30, 12, 15, 35, 10, 40])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # ファイル依存関係データ（スクリプト/Prefab/シーンのみ、アセットは別シート）
    files = [
        # ファイル名, 種別, 場所, 依存先, 状態, 備考
        ("【スクリプト】", "", "", "", "", ""),
        ("GameManager.cs", "Script", "Assets/Scripts", "", "未作成", "ゲーム全体の管理"),
        ("PaddleController.cs", "Script", "Assets/Scripts", "GameManager.cs", "未作成", "パドル操作"),
        ("BallController.cs", "Script", "Assets/Scripts", "GameManager.cs", "未作成", "ボール物理"),
        ("BrickController.cs", "Script", "Assets/Scripts", "GameManager.cs", "未作成", "ブロック破壊処理"),
        ("FishLayerManager.cs", "Script", "Assets/Scripts", "BrickController.cs", "未作成", "魚レイヤー表示管理"),
        ("UIManager.cs", "Script", "Assets/Scripts", "GameManager.cs", "未作成", "UI制御"),
        ("", "", "", "", "", ""),
        ("【Prefab】", "", "", "", "", ""),
        ("Brick.prefab", "Prefab", "Assets/Prefabs", "BrickController.cs", "未作成", "ブロックPrefab"),
        ("Ball.prefab", "Prefab", "Assets/Prefabs", "BallController.cs", "未作成", "ボールPrefab"),
        ("Paddle.prefab", "Prefab", "Assets/Prefabs", "PaddleController.cs", "未作成", "パドルPrefab"),
        ("", "", "", "", "", ""),
        ("【シーン】", "", "", "", "", ""),
        ("TitleScene.unity", "Scene", "Assets/Scenes", "", "未作成", "タイトル画面"),
        ("GameScene.unity", "Scene", "Assets/Scenes", "全Prefab", "未作成", "メインゲーム"),
        ("ResultScene.unity", "Scene", "Assets/Scenes", "", "未作成", "リザルト画面"),
    ]

    for row_idx, row_data in enumerate(files, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_data[0].startswith("【"):  # セクション見出し
                apply_section_style(cell)
            else:
                apply_normal_style(cell)
                # 状態に応じて色分け
                if col_idx == 5:  # 状態列
                    if value == "既存":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "要作成":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "未作成":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    return ws

# ============================================
# シート3: 画面遷移
# ============================================
def create_screen_flow_sheet(wb):
    ws = wb.create_sheet("03_画面遷移")

    headers = ["シーン", "状態", "表示要素", "操作", "遷移先", "備考"]
    set_column_widths(ws, [18, 15, 40, 30, 18, 35])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # 画面遷移データ
    flows = [
        ("TitleScene", "初期表示", "タイトルロゴ、スタートボタン", "スタートボタン押下", "GameScene", ""),
        ("", "", "", "", "", ""),
        ("GameScene", "Ready", "魚画像、グリッド、パドル、ボール（静止）", "スペース/クリック/タップ", "Playing", "ボール発射"),
        ("", "Playing", "スコア、残機、ゲーム進行中", "←→/マウス/タッチ", "", "パドル移動"),
        ("", "", "", "ブロック全破壊", "GameClear", ""),
        ("", "", "", "ボール落下（残機0）", "GameOver", ""),
        ("", "GameClear", "YOU WIN!表示、骨画像完全露出", "リスタートボタン", "Ready", ""),
        ("", "GameOver", "GAME OVER表示", "リスタートボタン", "Ready", ""),
        ("", "", "", "タイトルボタン", "TitleScene", ""),
        ("", "", "", "", "", ""),
        ("ResultScene", "（将来）", "ハイスコア表示など", "", "", "将来実装"),
    ]

    for row_idx, row_data in enumerate(flows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_data[0] and col_idx == 1:  # シーン名
                apply_section_style(cell)
            else:
                apply_normal_style(cell)

    return ws

# ============================================
# シート4: 画像アセット
# ============================================
def create_image_assets_sheet(wb):
    ws = wb.create_sheet("04_画像アセット")

    headers = ["ファイル名", "プレビュー", "サイズ", "用途", "使用箇所", "状態", "備考"]
    set_column_widths(ws, [25, 25, 12, 25, 20, 10, 30])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # 画像アセットデータ
    images = [
        # ファイル名, プレビュー, サイズ, 用途, 使用箇所, 状態, 備考
        ("【魚画像】", "", "", "", "", "", ""),
        ("sakana_normal.png", "", "─", "魚の絵（前面レイヤー）", "GameScene", "既存", "プレイヤーが描く"),
        ("sakana_bone.png", "", "─", "骨の絵（背面レイヤー）", "GameScene", "要作成", "プレイヤーが描く"),
        ("sakana_sliced.png", "", "─", "切り身の絵（中間）", "GameScene", "将来", "3段階演出用"),
        ("", "", "", "", "", "", ""),
        ("【UI画像】", "", "", "", "", "", ""),
        ("title_logo.png", "", "─", "タイトルロゴ", "TitleScene", "未作成", ""),
        ("button_start.png", "", "─", "スタートボタン", "TitleScene", "未作成", ""),
        ("button_restart.png", "", "─", "リスタートボタン", "GameScene", "未作成", ""),
        ("heart_icon.png", "", "─", "残機アイコン", "GameScene", "未作成", ""),
        ("", "", "", "", "", "", ""),
        ("【エフェクト】", "", "", "", "", "", ""),
        ("particle_break.png", "", "─", "ブロック破壊パーティクル", "GameScene", "未作成", ""),
        ("", "", "", "", "", "", ""),
        ("【背景】", "", "", "", "", "", ""),
        ("bg_title.png", "", "─", "タイトル背景", "TitleScene", "未作成", ""),
        ("bg_game.png", "", "─", "ゲーム背景", "GameScene", "未作成", ""),
    ]

    row_height = 80  # 画像用に行を高く
    for row_idx, row_data in enumerate(images, 2):
        if row_data[0].startswith("【") or row_data[0] == "":
            pass  # セクション行は通常の高さ
        else:
            ws.row_dimensions[row_idx].height = row_height

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_data[0].startswith("【"):
                apply_section_style(cell)
            else:
                apply_normal_style(cell)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 状態に応じて色分け
                if col_idx == 6:  # 状態列
                    if value == "既存":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "要作成":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "未作成":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 画像挿入を試みる
    image_path = os.path.join(ASSET_BASE, "sakana", "sakana_normal.png")
    if os.path.exists(image_path):
        try:
            img = XLImage(image_path)
            max_width = 120
            ratio = max_width / img.width
            img.width = max_width
            img.height = int(img.height * ratio)
            ws.add_image(img, "B3")
            ws.cell(row=3, column=3).value = f"{img.width}x{img.height}"
        except Exception as e:
            ws.cell(row=3, column=2).value = f"（エラー: {e}）"
    else:
        ws.cell(row=3, column=2).value = "（未検出）"

    return ws

# ============================================
# シート5: サウンドアセット
# ============================================
def create_sound_assets_sheet(wb):
    ws = wb.create_sheet("05_サウンドアセット")

    headers = ["ファイル名", "種別", "長さ", "用途", "再生タイミング", "状態", "備考"]
    set_column_widths(ws, [25, 10, 10, 25, 25, 10, 30])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # サウンドアセットデータ
    sounds = [
        # ファイル名, 種別, 長さ, 用途, 再生タイミング, 状態, 備考
        ("【BGM】", "", "", "", "", "", ""),
        ("bgm_title.mp3", "BGM", "─", "タイトルBGM", "TitleScene表示中", "未作成", "ループ再生"),
        ("bgm_game.mp3", "BGM", "─", "ゲームBGM", "GameScene Playing中", "未作成", "ループ再生"),
        ("bgm_clear.mp3", "BGM", "─", "クリアBGM", "GameClear時", "未作成", "ファンファーレ"),
        ("bgm_gameover.mp3", "BGM", "─", "ゲームオーバーBGM", "GameOver時", "未作成", ""),
        ("", "", "", "", "", "", ""),
        ("【SE（効果音）】", "", "", "", "", "", ""),
        ("se_ball_hit.wav", "SE", "─", "ボール反射音", "壁/パドル/ブロック衝突時", "未作成", ""),
        ("se_brick_break.wav", "SE", "─", "ブロック破壊音", "ブロック破壊時", "未作成", "魚をさばく音？"),
        ("se_ball_launch.wav", "SE", "─", "ボール発射音", "ボール発射時", "未作成", ""),
        ("se_life_lost.wav", "SE", "─", "残機減少音", "ボール落下時", "未作成", ""),
        ("se_button.wav", "SE", "─", "ボタン押下音", "UI操作時", "未作成", ""),
        ("se_score.wav", "SE", "─", "スコア加算音", "スコア増加時", "未作成", ""),
        ("", "", "", "", "", "", ""),
        ("【ボイス（将来）】", "", "", "", "", "", ""),
        ("voice_clear.wav", "Voice", "─", "クリアボイス", "GameClear時", "将来", ""),
        ("voice_gameover.wav", "Voice", "─", "ゲームオーバーボイス", "GameOver時", "将来", ""),
    ]

    for row_idx, row_data in enumerate(sounds, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_data[0].startswith("【"):
                apply_section_style(cell)
            else:
                apply_normal_style(cell)
                # 状態に応じて色分け
                if col_idx == 6:  # 状態列
                    if value == "既存":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "要作成":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "未作成":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    return ws

# ============================================
# メイン処理
# ============================================
def main():
    print("Sabake_osakana 仕様書を生成中...")

    wb = Workbook()

    # 各シート作成
    create_overview_sheet(wb)
    create_dependencies_sheet(wb)
    create_screen_flow_sheet(wb)
    create_image_assets_sheet(wb)
    create_sound_assets_sheet(wb)

    # 出力ディレクトリ確認
    output_dir = os.path.dirname(OUTPUT_PATH)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 保存
    wb.save(OUTPUT_PATH)
    print(f"✓ 仕様書を生成しました: {OUTPUT_PATH}")
    print("\nシート構成:")
    print("  01_全体概要       - プロジェクト基本情報")
    print("  02_ファイル依存関係 - スクリプト/Prefab/シーンの関係")
    print("  03_画面遷移       - シーン間の遷移フロー")
    print("  04_画像アセット    - 画像ファイル一覧（プレビュー付き）")
    print("  05_サウンドアセット - BGM/SE/ボイス一覧")

if __name__ == "__main__":
    main()
